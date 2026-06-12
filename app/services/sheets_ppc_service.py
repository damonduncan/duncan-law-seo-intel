"""Google Sheets PPC sync service.

Reads the PPC reporting spreadsheet (one tab per month) via the Google Drive
export API, parses any month tabs not yet in the database, and returns entries
ready to upsert into ppc_monthly_data.

Requires a refresh token with spreadsheets.readonly scope — obtained via the
one-time OAuth flow at /ppc/connect-google-sheets.
"""
import io
import logging
import re
from calendar import monthrange
from typing import Optional

logger = logging.getLogger(__name__)

_SPREADSHEET_ID = "1g01hXs3ECHCbbf2QiEl4-TY_d6gRZkQJgj9yC4cVYSQ"

_MONTH_MAP = {
    'january': 1, 'february': 2, 'march': 3, 'april': 4,
    'may': 5, 'june': 6, 'july': 7, 'august': 8,
    'september': 9, 'october': 10, 'november': 11, 'december': 12,
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4,
    'jun': 6, 'jul': 7, 'aug': 8, 'sep': 9,
    'oct': 10, 'nov': 11, 'dec': 12,
}
_MARKET_KEY = {
    'charlotte':     'charlotte',
    'salisbury':     'salisbury',
    'greensboro':    'greensboro',
    'winston-salem': 'winston_salem',
    'winston salem': 'winston_salem',
    'asheville':     'asheville',
}
_SKIP_TABS = {'ad spend since 2.4.23'}


# ── Credentials ───────────────────────────────────────────────────────────────

def load_sheets_refresh_token() -> Optional[str]:
    try:
        from app.database import SessionLocal
        from app.models.discovery import DiscoveryCache
        db = SessionLocal()
        try:
            row = db.query(DiscoveryCache).filter(
                DiscoveryCache.key == "google_sheets_ppc_credentials"
            ).first()
            if row and isinstance(row.value, dict):
                return row.value.get("refresh_token")
        finally:
            db.close()
    except Exception as e:
        logger.error(f"Sheets PPC: could not load refresh token: {e}")
    return None


def save_sheets_refresh_token(refresh_token: str) -> None:
    from datetime import datetime, timezone
    from app.database import SessionLocal
    from app.models.discovery import DiscoveryCache
    from app.models.base import new_uuid
    db = SessionLocal()
    try:
        payload = {
            "refresh_token": refresh_token,
            "saved_at": datetime.now(timezone.utc).isoformat(),
        }
        row = db.query(DiscoveryCache).filter(
            DiscoveryCache.key == "google_sheets_ppc_credentials"
        ).first()
        if row:
            row.value      = payload
            row.updated_at = datetime.now(timezone.utc)
        else:
            db.add(DiscoveryCache(
                id=new_uuid(),
                key="google_sheets_ppc_credentials",
                value=payload,
                updated_at=datetime.now(timezone.utc),
            ))
        db.commit()
    finally:
        db.close()


def is_configured() -> bool:
    from app.config import settings
    return bool(
        settings.google_client_id
        and settings.google_client_secret
        and load_sheets_refresh_token()
    )


def _get_access_token(client_id: str, client_secret: str, refresh_token: str) -> str:
    import httpx
    resp = httpx.post(
        "https://oauth2.googleapis.com/token",
        data={
            "grant_type":    "refresh_token",
            "client_id":     client_id,
            "client_secret": client_secret,
            "refresh_token": refresh_token,
        },
        timeout=15,
    )
    resp.raise_for_status()
    return resp.json()["access_token"]


# ── Parsing ───────────────────────────────────────────────────────────────────

def _parse_tab_name(name: str) -> tuple:
    parts = name.strip().split()
    if len(parts) == 2:
        m = _MONTH_MAP.get(parts[0].lower())
        if m and parts[1].isdigit():
            return int(parts[1]), m
    return None, None


def _clean_num(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace('$', '').replace(',', '').replace('%', '').strip()
    try:
        return float(s)
    except ValueError:
        return None


def _parse_worksheet(ws) -> Optional[dict]:
    """Extract market metrics from one openpyxl worksheet. Returns None if no data found."""
    rows = list(ws.iter_rows(values_only=True))

    header_row = None
    for i, row in enumerate(rows):
        if row[0] and str(row[0]).strip().lower() in ('google', 'market'):
            header_row = i
            break
    if header_row is None:
        return None

    entry = {}
    total_entry = None

    for row in rows[header_row + 1:]:
        if not row[0]:
            continue
        label = str(row[0]).strip().lower()
        if label in ('insights', 'moving forward', 'competitor report', 'notes'):
            break

        mkey = _MARKET_KEY.get(label)
        if not mkey and label != 'total':
            continue

        imp = _clean_num(row[1])
        clk = _clean_num(row[2])
        ctr = _clean_num(row[3])
        sp  = _clean_num(row[4])
        lds = _clean_num(row[6])
        cpl = _clean_num(row[7])

        # CTR stored as decimal in newer tabs (0.15 → 15%)
        if ctr is not None and ctr <= 1.0:
            ctr = round(ctr * 100, 2)
        elif ctr is not None:
            ctr = round(ctr, 2)

        d = {
            'impressions': int(imp) if imp else 0,
            'clicks':      int(clk) if clk else 0,
            'ctr':         ctr or 0.0,
            'spend':       round(sp, 2) if sp else 0.0,
            'leads':       int(lds) if lds else 0,
            'cpl':         round(cpl, 2) if cpl else 0.0,
        }

        if label == 'total':
            total_entry = d
        else:
            entry[mkey] = d

    if not entry:
        return None

    if total_entry:
        entry['total'] = total_entry
    else:
        mkts = ['charlotte', 'salisbury', 'greensboro', 'winston_salem', 'asheville']
        ti = tc = 0
        ts = tl = 0.0
        for m in mkts:
            if m in entry:
                ti += entry[m]['impressions']
                tc += entry[m]['clicks']
                ts += entry[m]['spend']
                tl += entry[m]['leads']
        tl_int = int(tl)
        entry['total'] = {
            'impressions': ti,
            'clicks':      tc,
            'ctr':         round(tc / ti * 100, 2) if ti else 0.0,
            'spend':       round(ts, 2),
            'leads':       tl_int,
            'cpl':         round(ts / tl_int, 2) if tl_int else 0.0,
        }

    return entry


# ── Main sync function ────────────────────────────────────────────────────────

def sync_new_months(db=None) -> dict:
    """Download the PPC spreadsheet, parse any new month tabs, upsert to DB.

    Returns {"imported": [(year, month), ...], "skipped": int, "errors": [...]}
    """
    import openpyxl
    import httpx
    from app.config import settings
    from app.models.discovery import DiscoveryCache
    from app.models.base import new_uuid
    from datetime import datetime, timezone

    refresh_token = load_sheets_refresh_token()
    if not refresh_token:
        return {"imported": [], "skipped": 0, "errors": ["Not configured — no refresh token"]}

    access_token = _get_access_token(
        settings.google_client_id,
        settings.google_client_secret,
        refresh_token,
    )

    # Download spreadsheet as Excel
    export_url = (
        f"https://www.googleapis.com/drive/v3/files/{_SPREADSHEET_ID}/export"
        f"?mimeType=application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp = httpx.get(
        export_url,
        headers={"Authorization": f"Bearer {access_token}"},
        timeout=60,
        follow_redirects=True,
    )
    resp.raise_for_status()

    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)

    # Determine which months are already in the DB
    close_db = False
    if db is None:
        from app.database import SessionLocal
        db = SessionLocal()
        close_db = True

    try:
        row = db.query(DiscoveryCache).filter(
            DiscoveryCache.key == "ppc_monthly_data"
        ).first()
        existing_raw = row.value if row else []
        if isinstance(existing_raw, str):
            import json as _j
            existing_raw = _j.loads(existing_raw)
        if isinstance(existing_raw, dict):
            existing_raw = existing_raw.get("months", [])

        existing_keys = {(m["year"], m["month"]) for m in existing_raw}

        imported = []
        skipped  = 0
        errors   = []
        new_entries = []

        for tab_name in wb.sheetnames:
            if tab_name.lower() in _SKIP_TABS:
                continue
            year, month = _parse_tab_name(tab_name)
            if not year:
                continue
            if (year, month) in existing_keys:
                skipped += 1
                continue

            try:
                ws = wb[tab_name]
                parsed = _parse_worksheet(ws)
                if parsed:
                    entry = {"year": year, "month": month, **parsed}
                    new_entries.append(entry)
                    imported.append((year, month))
                else:
                    errors.append(f"{tab_name}: no data rows found")
            except Exception as e:
                errors.append(f"{tab_name}: {e}")

        if new_entries:
            combined = existing_raw + new_entries
            combined.sort(key=lambda m: (m["year"], m["month"]))

            now = datetime.now(timezone.utc)
            if row:
                row.value      = combined
                row.updated_at = now
            else:
                db.add(DiscoveryCache(
                    id=new_uuid(), key="ppc_monthly_data",
                    value=combined, updated_at=now,
                ))

            # Record last sync timestamp
            sync_row = db.query(DiscoveryCache).filter(
                DiscoveryCache.key == "ppc_sheets_last_sync"
            ).first()
            if sync_row:
                sync_row.value      = {"synced_at": now.isoformat(), "imported": len(imported)}
                sync_row.updated_at = now
            else:
                db.add(DiscoveryCache(
                    id=new_uuid(), key="ppc_sheets_last_sync",
                    value={"synced_at": now.isoformat(), "imported": len(imported)},
                    updated_at=now,
                ))
            db.commit()

        logger.info(
            f"Sheets PPC sync: imported {len(imported)} new months, "
            f"skipped {skipped} existing, {len(errors)} errors"
        )
        return {"imported": imported, "skipped": skipped, "errors": errors}

    finally:
        if close_db:
            db.close()
